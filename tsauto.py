from openpyxl import load_workbook
wb = load_workbook(filename = 'tssample.xlsx')
ws = wb.active

# prompt user for abnormal hours
while True:
    answer = input("did you work less than 40 hours last week? ")
    if answer not in ["yes",'y','Y',"Yes","YES","no","n","No","NO"]:
        print('idk what you mean')
        continue
    else:
        break


if answer in ["yes",'y','Y',"Yes", "YES"]:
    m_hours = int(input("how many hours did you work on monday? "))
    if m_hours < 8:
        m_off = (input("which category do the other hours fall into? "))
        
        if m_off in ["sick"]:
            ws['B1'] = 8 - m_hours
        elif m_off in ["vacation"]:
            ws['C1'] = 8 - m_hours
        elif m_off in [""]:
            ws['C1'] = 8 - m_hours
        elif m_off in ["vacation"]:
            ws['C1'] = 8 - m_hours
        else:
            print("category not recognized")
        
    elif m_hours == 8:
        ws['A1'] = m_hours
    else:
        print("your input wasn't recognized")
        
    t_hours = int(input("how many hours did you work on tuesday? "))
    if t_hours < 8:
        t_off = (input("which category do the other hours fall into? "))
        
        if t_off in ["sick"]:
            ws['B1'] = 8 - t_hours
        elif t_off in ["vacation"]:
            ws['C1'] = 8 - t_hours
        elif t_off in [""]:
            ws['C1'] = 8 - t_hours
        elif t_off in ["vacation"]:
            ws['C1'] = 8 - t_hours
        else:
            print("category not recognized")
            
    elif t_hours == 8:
        ws['A1'] = t_hours
    else:
        print("your input wasn't recognized")
        
    w_hours = int(input("how many hours did you work on wednesday? "))
    if w_hours < 8:
        w_off = (input("which category do the other hours fall into? "))
        
        if w_off in ["sick"]:
            ws['B1'] = 8 - w_hours
        elif w_off in ["vacation"]:
            ws['C1'] = 8 - w_hours
        elif w_off in [""]:
            ws['C1'] = 8 - w_hours
        elif w_off in ["vacation"]:
            ws['C1'] = 8 - w_hours
        else:
            print("category not recognized")
    elif w_hours == 8:
        ws['A1'] = w_hours
    else:
        print("your input wasn't recognized")
        
    th_hours = int(input("how many hours did you work on thursday? "))
    if th_hours < 8:
        th_off = (input("which category do the other hours fall into? "))
        
        if th_off in ["sick"]:
            ws['B1'] = 8 - th_hours
        elif th_off in ["vacation"]:
            ws['C1'] = 8 - th_hours
        elif th_off in [""]:
            ws['C1'] = 8 - th_hours
        elif th_off in ["vacation"]:
            ws['C1'] = 8 - th_hours
        else:
            print("category not recognized")
    elif th_hours == 8:
        ws['A1'] = th_hours
    else:
        print("your input wasn't recognized")
        
    f_hours = int(input("how many hours did you work on friday? "))
    if f_hours < 8:
        f_off = (input("which category do the other hours fall into? "))
        
        if f_off in ["sick"]:
            ws['B1'] = 8 - f_hours
        elif f_off in ["vacation"]:
            ws['C1'] = 8 - f_hours
        elif f_off in [""]:
            ws['C1'] = 8 - f_hours
        elif f_off in ["vacation"]:
            ws['C1'] = 8 - f_hours
        else:
            print("category not recognized")
    elif f_hours == 8:
        ws['A1'] = f_hours
    else:
        print("your input wasn't recognized")
        
    ws['A1'] = m_hours
    ws['A2'] = t_hours
    ws['A3'] = w_hours
    ws['A4'] = th_hours
    ws['A4'] = f_hours
elif answer in ["no","NO","No","N","n"]:
    ws['A1'] = 8
    ws['A2'] = 8
    ws['A3'] = 8
    ws['A4'] = 8
    ws['A5'] = 8
else :
    print("wrong answer fucktard")

date = (input("enter todays date (mm/dd/yy)"))
pdate = (input("enter the day the period ends"))
         
ws['b9'] = date
ws['b10'] = pdate

wb.save('tssample.xlsx')

